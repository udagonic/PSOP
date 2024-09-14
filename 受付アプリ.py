import nfc
import openpyxl
from datetime import datetime

# 学生データ
# 出席学生データ
student_data = {
学籍番号：("学生名","所属専攻"),
}
# 欠席学生データ
absent_student_data = {
学籍番号：("学生名","所属専攻"),
}

# Excelファイルのパス
excel_file_path = r"【要変更】「参加受付表.xlsx」のファイルパス"


def extract_data(data_lines):
    result = []
    extracting = False

    for line in data_lines:
        if "Area 1A81--1AFF" in line:
            # 指定されたエリアに到達したらデータの抽出を開始
            extracting = True
            result.append(line)
        elif extracting and line.startswith("Random Service 106:"):
            # 指定されたサービスに到達したらデータを抽出
            result.append(line)
        elif extracting and line.startswith(" "):
            # スペースで始まるデータ行を抽出
            result.append(line)
        elif extracting:
            # 指定されたデータ以外の行に達したら抽出を終了
            break

    return result

def check_attendance(student_number):
    # 学生データと照合
    if student_number in student_data:
        student_name, department = student_data[student_number]
        print(f"受付確認しました。 所属専攻: {department}, 学生番号: {student_number}, 学生名: {student_name}")
        # 受付時間を取得
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"受付時間: {current_time}")
        # Excelファイルにデータを書き込む
        write_to_excel(student_number, student_name, department, current_time)
    elif student_number in absent_student_data:
        student_name, department = absent_student_data[student_number]
        print(f"受付できません。 所属専攻: {department}, 学生番号: {student_number}, 学生名: {student_name} 様は欠席で登録されています。")
    else:
        print(f"学生番号 {student_number} はデータに存在しません。")


def write_to_excel(student_number, student_name, department, attendance_time):
    try:
        # Excelファイルを開く（存在しない場合は新規作成）
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["所属専攻", "学生番号", "学生名", "受付時間"])

        sheet = workbook.active

        # 最終行を取得
        last_row = sheet.max_row + 1

        # データを書き込む
        sheet.cell(row=last_row, column=1, value=department)
        sheet.cell(row=last_row, column=2, value=student_number)
        sheet.cell(row=last_row, column=3, value=student_name)
        sheet.cell(row=last_row, column=4, value=attendance_time)

        # 保存
        workbook.save(excel_file_path)
        print(f"Excelファイルにデータを書き込みました。")
    except Exception as e:
        print(f"Excelファイルへの書き込み中にエラーが発生しました: {e}")


def on_connect(tag: nfc.tag.Tag) -> bool:
    try:
        print("connected")
        data_lines = tag.dump()
        extracted_data = extract_data(data_lines)
        
        if len(extracted_data) >= 3:
            # 3行目のデータを抽出
            extracted_line = extracted_data[2]
            
            # '|' で囲まれた数字のうち、3文字目から10文字目までを抜き出す
            start_index = extracted_line.find("|") + 3
            end_index = start_index + 10
            extracted_numbers = extracted_line[start_index:end_index]
            
            print(extracted_numbers)
            
            # 出力した数字を学生データと照合
            check_attendance(int(extracted_numbers))
    except Exception as e:
        print(f"エラーが発生しました: {e}")
    
    return True

def on_release(tag: nfc.tag.Tag) -> None:
    print("released")

with nfc.ContactlessFrontend("usb") as clf:
    while True:
        clf.connect(rdwr={"on-connect": on_connect, "on-release": on_release})

